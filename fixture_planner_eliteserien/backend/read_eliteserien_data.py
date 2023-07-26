from constants import python_anywhere_path, total_number_of_eliteserien_teams, total_number_of_eliteserien_teams, total_number_of_gameweeks_in_eliteserien, fixture_folder_name, stored_data, esf, current_season_name_eliteserien
from models.fixtures.models.TeamFixtureInfoEliteserienModel import TeamFixtureInfoEliteserienModel
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook


def read_eliteserien_excel_to_db_format(defensivt=""):
    xlsx_file = "Eliteserien_fixtures" + defensivt + ".xlsx"
    
    path = python_anywhere_path + stored_data + "/" + esf + "/" + current_season_name_eliteserien + \
        "/" + fixture_folder_name + "/" + xlsx_file

    max_teams = total_number_of_eliteserien_teams
    max_games = total_number_of_gameweeks_in_eliteserien + 2
    
    wb = load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    ws_fdr = wb[sheet_names[0]]

    color_to_fdr_dict = get_fdr_colors_to_difficulty_rating_dict(ws_fdr)
    column_letters_list = get_list_of_column_letters_from_int_range(2, max_games)
    row_numbers_list = get_list_of_row_numbers_from_int_range(3, max_teams + 3)

    fixture_info_list, team_name_hexcolor = [], []

    max_gws, team_id_idx = 0, 1

    for row in row_numbers_list:
        temp_oppTeamNameList, temp_oppTeamHomeAwayList, temp_oppTeamDifficultyScore, temp_gw = [], [], [], []
        
        team_name_cell_name = "{}{}".format('A', row)
        team_name = ws_fdr[team_name_cell_name].value.split("(")[0]
        team_name_short = ws_fdr[team_name_cell_name].value.split("(")[1][:-1]
        team_color_hex = "0" if ws_fdr[team_name_cell_name].fill.start_color.index == "00000000" else  ws_fdr[team_name_cell_name].fill.start_color.index
        team_font_color_hex = ws_fdr[team_name_cell_name].font.color.index
        
        team_name_hexcolor.append([team_name, team_color_hex, team_font_color_hex, team_id_idx])

        for column in column_letters_list:
            cell_name = "{}{}".format(column, row)
            gw_number = "{}{}".format(column, 2)
            value = ws_fdr[cell_name].value
           
            if value is not None:
                all_games_in_gw = value.split("+")
                for game in all_games_in_gw:
                    fdr = color_to_fdr_dict[ws_fdr[cell_name].fill.start_color.index]
                    opponent = game
                    home_away = "A" if game.islower() else "H"
                    temp_oppTeamNameList.append(opponent)
                    temp_oppTeamHomeAwayList.append(home_away)
                    temp_oppTeamDifficultyScore.append(fdr)
                    temp_gw.append(ws_fdr[gw_number].value)
        
        max_gws = max(temp_gw)
        
        fixture_info_list.append(TeamFixtureInfoEliteserienModel(team_name=team_name,
            team_id=team_id_idx, team_short_name=team_name_short, date="",
            opp_team_name_list=temp_oppTeamNameList,
            opp_team_home_away_list=temp_oppTeamHomeAwayList,
            opp_team_difficulty_score=temp_oppTeamDifficultyScore,
            gw=temp_gw)
        )
        
        team_id_idx += 1

    dates = [gw_i for gw_i in range(max_gws)]

    fdr_to_colors_dict = get_difficulty_rating_to_fdr_colors_dict(ws_fdr)

    return fixture_info_list, dates, fdr_to_colors_dict, team_name_hexcolor
   

def get_fdr_colors_to_difficulty_rating_dict(data, DG_score=0.5, blank_score=10):
    fdr_1 = data['B1'].fill.start_color.index
    fdr_2 = data['C1'].fill.start_color.index
    fdr_3 = data['D1'].fill.start_color.index
    fdr_4 = data['E1'].fill.start_color.index
    fdr_5 = data['F1'].fill.start_color.index
    fdr_dobbel = data['H1'].fill.start_color.index

    # if all are equal, then set the fdr rating to 0 (not given)
    if (fdr_1 == fdr_2 == fdr_3 ==  fdr_4 == fdr_5 == fdr_dobbel):
        return {
            fdr_1: 0,
        }
    
    return {
        fdr_dobbel: DG_score,
        fdr_1: 1,
        fdr_2: 2,
        fdr_3: 3,
        fdr_4: 4,
        fdr_5: 5,
        "Not implemented": blank_score,
    }


def get_difficulty_rating_to_fdr_colors_dict(data, DG_score=0.5, blank_score=10):
    return {
        0: data['I1'].fill.start_color.index,
        DG_score: data['H1'].fill.start_color.index,
        1: data['B1'].fill.start_color.index,
        2: data['C1'].fill.start_color.index,
        3: data['D1'].fill.start_color.index,
        4: data['E1'].fill.start_color.index,
        5: data['F1'].fill.start_color.index,
        blank_score: "Not implemented"
    }


def get_list_of_column_letters_from_int_range(start, stop):
    return [get_column_letter(idx) for idx in range(start, stop)]


def get_list_of_row_numbers_from_int_range(start, stop):
    return [idx for idx in range(start, stop)]